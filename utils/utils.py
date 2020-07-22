from builtins import print
import numpy as np
import pandas as pd
import matplotlib
import pickle
from augmentation.src import dtw_mean as DTW
from datetime import datetime
matplotlib.use('agg')
import matplotlib.pyplot as plt
from openpyxl.styles import  PatternFill, Font
matplotlib.rcParams['font.family'] = 'sans-serif'
matplotlib.rcParams['font.sans-serif'] = 'Arial'
import os
import operator
import utils
from utils.constants import UNIVARIATE_DATASET_NAMES as DATASET_NAMES
from utils.constants import UNIVARIATE_DATASET_NAMES_2018 as DATASET_NAMES_2018
from utils.constants import ARCHIVE_NAMES  as ARCHIVE_NAMES
from utils.constants import CLASSIFIERS
from utils.constants import ITERATIONS
from utils.constants import ROOT_DIR
from utils.constants import MTS_DATASET_NAMES
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from openpyxl import load_workbook
from openpyxl import Workbook
from scipy.interpolate import interp1d
from scipy.interpolate import make_interp_spline
from scipy.io import loadmat


def generate_save_paths(X_data, y_data, index, path_to_file):
    '''
    function to generate and save the warping paths
     - X_data: The data to generate the paths
     - y_data: The labels according to each time series
     - index: The index according to each time series in X_data
     - path_to_file: String with indicates where the file for the paths and their indices will be saved
     
    '''
    if os.path.isfile(path_to_file + "_paths.txt") and os.path.isfile(path_to_file + "path_indices.txt"):
        print('paths already generated')
        return
    print('generate and save paths')
    paths = []
    index_combination = []

    classes = np.unique(y_data)

    length = X_data.shape[1]
    for c in classes: 
        ind = index[y_data == c]
        for i in ind:
            for j in ind:
                if i == j:
                    continue
                d, path = DTW.dtw(np.reshape(X_data[i,:], (length,1)),np.reshape(X_data[j,:], (length,1)), path = True)
                paths.append(path)
                index_combination.append([c,i,j])
        print(f"Generated paths for class {c}")


    with open(path_to_file + "_paths.txt", "wb") as f:
        pickle.dump(paths, f)
    
    with open(path_to_file + "path_indices.txt", "wb") as f:
        pickle.dump(index_combination, f)
        
    print('done')

def color_the_best_metric(excelfile,sheetname,best_metrics):
    wb = load_workbook(excelfile)
    ws = wb[sheetname]
   
    ## Create a dictionary of column names
    ColNames = {}
    Current  = 0
    for COL in ws.iter_cols(1, ws.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    ## Now you can access by column name
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for columnname in best_metrics.keys():
            if type(row_cells[ColNames[columnname]].value) is str:
                if float(row_cells[ColNames[columnname]].value.split('(')[0])==best_metrics[columnname]:
                    
                    row_cells[ColNames[columnname]].fill = PatternFill("solid", fgColor="DDDDDD")
                    row_cells[ColNames[columnname]].font = Font(b=True)
                       
    wb.save(excelfile)
            
def copy_excel(source,target):
    wb = load_workbook(source)
    wb.save(target)
    
def copy_eval_results_and_check_best_models():
    #copy the evaluation results
    source=ROOT_DIR+'/results/results.xlsx'
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M")
    target=source[:-5]+'_Copy'+dt_string +'.xlsx'
    
    copy_excel(source,target)
    
    #check best models and color the best metrics
    best_metrics={}
    datasets=DATASET_NAMES_2018
    df1=pd.read_excel(target,'results')
    
    for d in datasets:    
        col=df1[d]
        
        metrics=[]
        for metric in col:
            if type(metric) is str: 
                metric0=float(metric.split('(')[0])
                metrics.append(metric0)
                
        if (len(metrics)!=0):
            best_metrics.update({d:max(metrics)})
        
    color_the_best_metric(target,'results',best_metrics )

def plot_epochs_metric1(data, aug_hists, file_name, loc, metric='loss'):
    file_name=file_name+metric+'.png'
    plt.figure()
    for aug in aug_hists.keys():
        #smoothing
        #plot every 20 epochs
        # y=aug_hists[aug][metric][::20]
        # x=np.arange(start=0, stop=len(aug_hists[aug][metric]), step=1)[::20]
        # x_new = np.linspace(1, 2000, 500)
        # a_BSpline = make_interp_spline(x, y)
        # y_new = a_BSpline(x_new)
        # plt.plot(x_new, y_new)
        
        #without smoothing
        plt.plot(aug_hists[aug][metric][::20])
    
    plt.legend(aug_hists.keys(), loc='best')
    
    plt.title(metric+'_dataset:' + data)
    plt.ylabel(metric, fontsize='large')
    plt.xlabel('epoch', fontsize='large')
   
    plt.savefig(file_name, bbox_inches='tight')
    plt.close()      
  
###########
def plot_epochs_overview():
    
    cls_dirs= [d for d in os.listdir(ROOT_DIR+'/results') if os.path.isdir(ROOT_DIR+'/results/'+d)]
 
    
    for cls_dir in cls_dirs:
        
       
        dir0 = ROOT_DIR+'/results/'+cls_dir
        appr_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
       
        for appr_dir in appr_dirs:
            
            
            dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir
            aug_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
            
            aug_dataset_dict={}
            
            for aug_dir in aug_dirs:
                
                #result['augmentation']=aug_dir
                dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir
                data_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
                
                dataset_hist_dict = {}
                for data_dir in data_dirs:
                    
                     if data_dir in DATASET_NAMES_2018:
                         dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir+'/'+data_dir
                         sub_dirs = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
                         
                         dfs=[]
                         if 'DONE' in sub_dirs:
                             for sub_dir in sub_dirs:
                                 dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir+'/'+data_dir+'/'+sub_dir
                                 #print(dir0)
                                 if os.path.isfile(dir0+'/history.csv'):
                                     df = pd.read_csv(dir0+'/history.csv')
                                     
                                     if 'acc' in list(df):
                            
                                         df = df.rename({'acc': 'accuracy'},axis=1)
                                     if 'val_acc' in list(df):
                                         
                                         df = df.rename({'val_acc': 'val_accuracy'},axis=1)
                                     dfs.append(df)
                             
                             avg_df=df
                             avg_df['loss']=(dfs[0] ['loss'] +  dfs[1] ['loss'] + dfs[2] ['loss'] + dfs[3] ['loss'] )/ 4 
                             avg_df['val_loss']=(dfs[0] ['val_loss'] +  dfs[1] ['val_loss'] + dfs[2] ['val_loss'] + dfs[3] ['val_loss'] )/ 4 
                             avg_df['accuracy']=(dfs[0] ['accuracy'] +  dfs[1] ['accuracy'] + dfs[2] ['accuracy'] + dfs[3] ['accuracy'] )/ 4 
                             avg_df['val_accuracy']=(dfs[0] ['val_accuracy'] +  dfs[1] ['val_accuracy'] + dfs[2] ['val_accuracy'] + dfs[3] ['val_accuracy'] )/ 4
                             dataset_hist_dict.update({data_dir:avg_df})
                                
               
                if aug_dir=='Random_diag0.5_down0.25_right0.25_each1_n1':
                    aug_dir='Random_n1'#
                if aug_dir=='Random_diag0.5_down0.25_right0.25_each1_n10':
                    aug_dir='Random_n10'#
                aug_dataset_dict.update({aug_dir:dataset_hist_dict})
                
           
            for data in DATASET_NAMES_2018:
                aug_hists={}
               
                for aug in aug_dirs:
                    if aug=='Random_diag0.5_down0.25_right0.25_each1_n1' :#
                       aug='Random_n1'#
                    if aug=='Random_diag0.5_down0.25_right0.25_each1_n10' :#
                       aug='Random_n10'#
                    dataset_hist=aug_dataset_dict[aug]
                    if data in dataset_hist.keys():
                        hist=dataset_hist[data]
                        aug_hists.update({aug:hist})
                
                if len(aug_hists)!=0:
                    
                    file_name=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+data+'_epochs_'
                    plot_epochs_metric1(data, aug_hists, file_name, 'upper left',metric='loss')
                    plot_epochs_metric1(data, aug_hists, file_name, 'upper left',metric='val_loss')
                    plot_epochs_metric1(data, aug_hists, file_name, 'lower right',metric='accuracy')
                    plot_epochs_metric1(data, aug_hists, file_name, 'lower right',metric='val_accuracy')
                     
    
def generate_results_overview():
    df_cols=['approach','classifier','augmentation']
    df_cols = df_cols +DATASET_NAMES_2018
             
    rows=[]
    dfs=[]
    
    result = {'approach':None,'classifier': None,'augmentation':None}
    for data in DATASET_NAMES_2018:
        result.update({data:None})
              
    
    cls_dirs= [d for d in os.listdir(ROOT_DIR+'/results') if os.path.isdir(ROOT_DIR+'/results/'+d)]
 
    
    for cls_dir in cls_dirs:
        
        result['classifier']=cls_dir
        dir0 = ROOT_DIR+'/results/'+cls_dir
        appr_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
       
        for appr_dir in appr_dirs:
            
            result['approach']=appr_dir
            dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir
            aug_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
            
            for aug_dir in aug_dirs:
                
                result['augmentation']=aug_dir
                dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir
                data_dirs  = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
                
                for data_dir in data_dirs:
                    
                     dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir+'/'+data_dir
                     sub_dirs = [d for d in os.listdir(dir0) if os.path.isdir(dir0+'/'+d)]
                     
                     res=[]
                     if 'DONE' in sub_dirs:
                         for sub_dir in sub_dirs:
                             
                                dir0=ROOT_DIR+'/results/'+cls_dir+'/'+appr_dir+'/'+aug_dir+'/'+data_dir+'/'+sub_dir
                                if os.path.isfile(dir0+'/df_metrics.csv'):
                                    
                                    df=pd.read_csv(dir0+'/df_metrics.csv')
                                    acc = df['accuracy'][0] 
                                    res.append(acc)
                                   
                         if len(res)!=0:
                        
                            res=np.array(res)                            
                            mean=round(np.mean(res)*100,1)
                            std=round(np.std(res)*100,1)
                            result[data_dir]=str(mean)+'('+str(std)+')'
   
                rows.append(result)
                df = pd.DataFrame(rows, columns = df_cols)                
                dfs.append(df)
                rows=[]
                for data_dir in data_dirs:
                    result[data_dir]=None
               
    df = pd.concat(dfs)  
    print(df)
    append_eval_results_to_excel(ROOT_DIR+'/results/'+'results.xlsx', 'results', df)

def append_eval_results_to_excel(excelfile,sheetname,df):
    if not os.path.isfile(excelfile):
        print('The excel file is not existing, creating a new excel file...'+excelfile)       
        wb = Workbook()
        wb.save(excelfile)
        
        
    wb = load_workbook(excelfile)
    if not (sheetname in wb.sheetnames):
        print('The worksheet is not existing, creating a new worksheet...'+sheetname) 
        ws1 = wb.create_sheet(sheetname)
        ws1.title = sheetname
        wb.save(excelfile)
    
            
    book = load_workbook(excelfile)
    idx = wb.sheetnames.index(sheetname)
    ws=book.get_sheet_by_name(sheetname)
    book.remove(ws)
    book.create_sheet(sheetname, idx)
    writer = pd.ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    df.to_excel(writer,sheet_name=sheetname, index = False,header= True)
    writer.save()

def check_dir(dir):
    done=dir+'/DONE'
    if os.path.isdir(dir):
        if os.path.isdir(done):
            
            print ('Model already exisits: ',dir[len(ROOT_DIR):])
            return True
    else:
        #print('Model not exisits: ',dir)            
        return False
def readucr(filename):
    data = np.loadtxt(filename, delimiter=',')
    Y = data[:, 0]
    X = data[:, 1:]
    return X, Y


def create_directory(directory_path):
    if os.path.exists(directory_path):
        return None
    else:
        try:
            os.makedirs(directory_path)
        except:
            # in case another machine created the path meanwhile !:(
            return None
        return directory_path


def create_path(root_dir, classifier_name, archive_name):
    output_directory = root_dir + '/results/' + classifier_name + '/' + archive_name + '/'
    if os.path.exists(output_directory):
        return None
    else:
        os.makedirs(output_directory)
        return output_directory


def read_dataset(root_dir, archive_name, dataset_name):
    datasets_dict = {}
    cur_root_dir = root_dir.replace('-temp', '')

    if archive_name == 'mts_archive':
        file_name = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name + '/'
        x_train = np.load(file_name + 'x_train.npy')
        y_train = np.load(file_name + 'y_train.npy')
        x_test = np.load(file_name + 'x_test.npy')
        y_test = np.load(file_name + 'y_test.npy')

        datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                       y_test.copy())

    elif archive_name == 'UCRArchive_2018':
        root_dir_dataset = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name 
        #+ '/'
        df_train = pd.read_csv(root_dir_dataset + '/' + dataset_name + '_TRAIN.tsv', sep='\t', header=None)

        df_test = pd.read_csv(root_dir_dataset + '/' + dataset_name + '_TEST.tsv', sep='\t', header=None)

        y_train = df_train.values[:, 0]
        y_test = df_test.values[:, 0]

        x_train = df_train.drop(columns=[0])
        x_test = df_test.drop(columns=[0])

        x_train.columns = range(x_train.shape[1])
        x_test.columns = range(x_test.shape[1])

        x_train = x_train.values
        x_test = x_test.values

        # znorm
        std_ = x_train.std(axis=1, keepdims=True)
        std_[std_ == 0] = 1.0
        x_train = (x_train - x_train.mean(axis=1, keepdims=True)) / std_

        std_ = x_test.std(axis=1, keepdims=True)
        std_[std_ == 0] = 1.0
        x_test = (x_test - x_test.mean(axis=1, keepdims=True)) / std_

        datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                       y_test.copy())
    else:
        file_name = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name + '/' + dataset_name
        x_train, y_train = readucr(file_name + '_TRAIN')
        x_test, y_test = readucr(file_name + '_TEST')
        datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                       y_test.copy())

    return datasets_dict


def read_all_datasets(root_dir, archive_name, split_val=False):
    datasets_dict = {}
    cur_root_dir = root_dir.replace('-temp', '')
    dataset_names_to_sort = []

    if archive_name == 'mts_archive':

        for dataset_name in MTS_DATASET_NAMES:
            root_dir_dataset = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name + '/'

            x_train = np.load(root_dir_dataset + 'x_train.npy')
            y_train = np.load(root_dir_dataset + 'y_train.npy')
            x_test = np.load(root_dir_dataset + 'x_test.npy')
            y_test = np.load(root_dir_dataset + 'y_test.npy')

            datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                           y_test.copy())
    elif archive_name == 'UCRArchive_2018':
        for dataset_name in DATASET_NAMES_2018:
            root_dir_dataset = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name + '/'

            df_train = pd.read_csv(root_dir_dataset + '/' + dataset_name + '_TRAIN.tsv', sep='\t', header=None)

            df_test = pd.read_csv(root_dir_dataset + '/' + dataset_name + '_TEST.tsv', sep='\t', header=None)

            y_train = df_train.values[:, 0]
            y_test = df_test.values[:, 0]

            x_train = df_train.drop(columns=[0])
            x_test = df_test.drop(columns=[0])

            x_train.columns = range(x_train.shape[1])
            x_test.columns = range(x_test.shape[1])

            x_train = x_train.values
            x_test = x_test.values

            # znorm
            std_ = x_train.std(axis=1, keepdims=True)
            std_[std_ == 0] = 1.0
            x_train = (x_train - x_train.mean(axis=1, keepdims=True)) / std_

            std_ = x_test.std(axis=1, keepdims=True)
            std_[std_ == 0] = 1.0
            x_test = (x_test - x_test.mean(axis=1, keepdims=True)) / std_

            datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                           y_test.copy())

    else:
        for dataset_name in DATASET_NAMES:
            root_dir_dataset = cur_root_dir + '/archives/' + archive_name + '/' + dataset_name + '/'
            file_name = root_dir_dataset + dataset_name
            x_train, y_train = readucr(file_name + '_TRAIN')
            x_test, y_test = readucr(file_name + '_TEST')

            datasets_dict[dataset_name] = (x_train.copy(), y_train.copy(), x_test.copy(),
                                           y_test.copy())

            dataset_names_to_sort.append((dataset_name, len(x_train)))

        dataset_names_to_sort.sort(key=operator.itemgetter(1))

        for i in range(len(DATASET_NAMES)):
            DATASET_NAMES[i] = dataset_names_to_sort[i][0]

    return datasets_dict


def get_func_length(x_train, x_test, func):
    if func == min:
        func_length = np.inf
    else:
        func_length = 0

    n = x_train.shape[0]
    for i in range(n):
        func_length = func(func_length, x_train[i].shape[1])

    n = x_test.shape[0]
    for i in range(n):
        func_length = func(func_length, x_test[i].shape[1])

    return func_length


def transform_to_same_length(x, n_var, max_length):
    n = x.shape[0]

    # the new set in ucr form np array
    ucr_x = np.zeros((n, max_length, n_var), dtype=np.float64)

    # loop through each time series
    for i in range(n):
        mts = x[i]
        curr_length = mts.shape[1]
        idx = np.array(range(curr_length))
        idx_new = np.linspace(0, idx.max(), max_length)
        for j in range(n_var):
            ts = mts[j]
            # linear interpolation
            f = interp1d(idx, ts, kind='cubic')
            new_ts = f(idx_new)
            ucr_x[i, :, j] = new_ts

    return ucr_x


def transform_mts_to_ucr_format():
    mts_root_dir = '/mnt/Other/mtsdata/'
    mts_out_dir = '/mnt/nfs/casimir/archives/mts_archive/'
    for dataset_name in MTS_DATASET_NAMES:
        # print('dataset_name',dataset_name)

        out_dir = mts_out_dir + dataset_name + '/'

        # if create_directory(out_dir) is None:
        #     print('Already_done')
        #     continue

        a = loadmat(mts_root_dir + dataset_name + '/' + dataset_name + '.mat')
        a = a['mts']
        a = a[0, 0]

        dt = a.dtype.names
        dt = list(dt)

        for i in range(len(dt)):
            if dt[i] == 'train':
                x_train = a[i].reshape(max(a[i].shape))
            elif dt[i] == 'test':
                x_test = a[i].reshape(max(a[i].shape))
            elif dt[i] == 'trainlabels':
                y_train = a[i].reshape(max(a[i].shape))
            elif dt[i] == 'testlabels':
                y_test = a[i].reshape(max(a[i].shape))

        # x_train = a[1][0]
        # y_train = a[0][:,0]
        # x_test = a[3][0]
        # y_test = a[2][:,0]

        n_var = x_train[0].shape[0]

        max_length = get_func_length(x_train, x_test, func=max)
        min_length = get_func_length(x_train, x_test, func=min)

        print(dataset_name, 'max', max_length, 'min', min_length)
        print()
        # continue

        x_train = transform_to_same_length(x_train, n_var, max_length)
        x_test = transform_to_same_length(x_test, n_var, max_length)

        # save them
        np.save(out_dir + 'x_train.npy', x_train)
        np.save(out_dir + 'y_train.npy', y_train)
        np.save(out_dir + 'x_test.npy', x_test)
        np.save(out_dir + 'y_test.npy', y_test)

        print('Done')


def calculate_metrics(y_true, y_pred, duration,y_true_val=None, y_pred_val=None):
    res = pd.DataFrame(data=np.zeros((1, 4), dtype=np.float), index=[0],
                       columns=['precision', 'accuracy', 'recall', 'duration'])
    res['precision'] = precision_score(y_true, y_pred, average='macro',zero_division=0)
    res['accuracy'] = accuracy_score(y_true, y_pred)

    if not y_true_val is None:
        # this is useful when transfer learning is used with cross validation
        res['accuracy_val'] = accuracy_score(y_true_val, y_pred_val)

    res['recall'] = recall_score(y_true, y_pred, average='macro')
    res['duration'] = duration
    #res['aug_duration'] = aug_time
    return res

def calculate_metrics1(y_true, y_pred, duration, aug_time,original,augmentated,y_true_val=None, y_pred_val=None):
    res = pd.DataFrame(data=np.zeros((1, 6), dtype=np.float), index=[0],
                       columns=['precision', 'accuracy', 'recall', 'duration','aug_duration','aug_factor'])
    res['precision'] = precision_score(y_true, y_pred, average='macro',zero_division=0)
    res['accuracy'] = accuracy_score(y_true, y_pred)

    if not y_true_val is None:
        # this is useful when transfer learning is used with cross validation
        res['accuracy_val'] = accuracy_score(y_true_val, y_pred_val)

    res['recall'] = recall_score(y_true, y_pred, average='macro')
    res['duration'] = duration
    res['aug_duration'] = aug_time
    #res['original training data size'] = original
    res['aug_factor'] = round(augmentated/original,2)
    return res

def save_test_duration(file_name, test_duration):
    res = pd.DataFrame(data=np.zeros((1, 1), dtype=np.float), index=[0],
                       columns=['test_duration'])
    res['test_duration'] = test_duration
    res.to_csv(file_name, index=False)


def generate_results_csv(output_file_name, root_dir):
    res = pd.DataFrame(data=np.zeros((0, 7), dtype=np.float), index=[],
                       columns=['classifier_name', 'archive_name', 'dataset_name',
                                'precision', 'accuracy', 'recall', 'duration'])
    for classifier_name in CLASSIFIERS:
        for archive_name in ARCHIVE_NAMES:
            datasets_dict = read_all_datasets(root_dir, archive_name)
            for it in range(ITERATIONS):
                curr_archive_name = archive_name
                if it != 0:
                    curr_archive_name = curr_archive_name + '_itr_' + str(it)
                for dataset_name in datasets_dict.keys():
                    output_dir = root_dir + '/results/' + classifier_name + '/' \
                                 + curr_archive_name + '/' + dataset_name + '/' + 'df_metrics.csv'
                    if not os.path.exists(output_dir):
                        continue
                    df_metrics = pd.read_csv(output_dir)
                    df_metrics['classifier_name'] = classifier_name
                    df_metrics['archive_name'] = archive_name
                    df_metrics['dataset_name'] = dataset_name
                    res = pd.concat((res, df_metrics), axis=0, sort=False)

    res.to_csv(root_dir + output_file_name, index=False)
    # aggreagte the accuracy for iterations on same dataset
    res = pd.DataFrame({
        'accuracy': res.groupby(
            ['classifier_name', 'archive_name', 'dataset_name'])['accuracy'].mean()
    }).reset_index()

    return res


def plot_epochs_metric(hist, file_name, metric='loss'):
    plt.figure()
    plt.plot(hist.history[metric])
    plt.plot(hist.history['val_' + metric])
    plt.title('model ' + metric)
    plt.ylabel(metric, fontsize='large')
    plt.xlabel('epoch', fontsize='large')
    plt.legend(['train', 'val'], loc='upper left')
    plt.savefig(file_name, bbox_inches='tight')
    plt.close()


def save_logs_t_leNet(output_directory, hist, y_pred, y_true, duration):
    hist_df = pd.DataFrame(hist.history)
    hist_df.to_csv(output_directory + 'history.csv', index=False)

    df_metrics = calculate_metrics(y_true, y_pred, duration)
    df_metrics.to_csv(output_directory + 'df_metrics.csv', index=False)

    index_best_model = hist_df['loss'].idxmin()
    row_best_model = hist_df.loc[index_best_model]

    df_best_model = pd.DataFrame(data=np.zeros((1, 6), dtype=np.float), index=[0],
                                 columns=['best_model_train_loss', 'best_model_val_loss', 'best_model_train_acc',
                                          'best_model_val_acc', 'best_model_learning_rate', 'best_model_nb_epoch'])

    df_best_model['best_model_train_loss'] = row_best_model['loss']
    df_best_model['best_model_val_loss'] = row_best_model['val_loss']
    df_best_model['best_model_train_acc'] = row_best_model['acc']
    df_best_model['best_model_val_acc'] = row_best_model['val_acc']
    df_best_model['best_model_nb_epoch'] = index_best_model

    df_best_model.to_csv(output_directory + 'df_best_model.csv', index=False)

    # plot losses
    plot_epochs_metric(hist, output_directory + 'epochs_loss.png')


def save_logs(output_directory, hist, y_pred, y_true, duration, lr=True, y_true_val=None, y_pred_val=None):
    hist_df = pd.DataFrame(hist.history)
    hist_df.to_csv(output_directory + 'history.csv', index=False)

    df_metrics = calculate_metrics(y_true, y_pred, duration, y_true_val, y_pred_val)
    df_metrics.to_csv(output_directory + 'df_metrics.csv', index=False)

    index_best_model = hist_df['loss'].idxmin()
    row_best_model = hist_df.loc[index_best_model]

    df_best_model = pd.DataFrame(data=np.zeros((1, 6), dtype=np.float), index=[0],
                                 columns=['best_model_train_loss', 'best_model_val_loss', 'best_model_train_acc',
                                          'best_model_val_acc', 'best_model_learning_rate', 'best_model_nb_epoch'])

    df_best_model['best_model_train_loss'] = row_best_model['loss']
    df_best_model['best_model_val_loss'] = row_best_model['val_loss']
    df_best_model['best_model_train_acc'] = row_best_model['accuracy']
    df_best_model['best_model_val_acc'] = row_best_model['val_accuracy']
    if lr == True:
        df_best_model['best_model_learning_rate'] = row_best_model['lr']
    df_best_model['best_model_nb_epoch'] = index_best_model

    df_best_model.to_csv(output_directory + 'df_best_model.csv', index=False)

    # for FCN there is no hyperparameters fine tuning - everything is static in code

    # plot losses
    plot_epochs_metric(hist, output_directory + 'epochs_loss.png')

    return df_metrics

# def save_logs_cv(output_directory, y_pred, y_true, duration):
    

#     df_metrics = calculate_metrics(y_true, y_pred, duration, y_true_val, y_pred_val)
#     df_metrics.to_csv(output_directory + 'CV_df_metrics.csv', index=False)

#     return df_metrics


def visualize_filter(root_dir):
    import tensorflow.keras as keras
    classifier = 'resnet'
    archive_name = 'UCRArchive_2018'
    dataset_name = 'GunPoint'
    datasets_dict = read_dataset(root_dir, archive_name, dataset_name)

    x_train = datasets_dict[dataset_name][0]
    y_train = datasets_dict[dataset_name][1]

    x_train = x_train.reshape(x_train.shape[0], x_train.shape[1], 1)

    model = keras.models.load_model(
        root_dir + 'results/' + classifier + '/' + archive_name + '/' + dataset_name + '/best_model.hdf5')

    # filters
    filters = model.layers[1].get_weights()[0]

    new_input_layer = model.inputs
    new_output_layer = [model.layers[1].output]

    new_feed_forward = keras.backend.function(new_input_layer, new_output_layer)

    classes = np.unique(y_train)

    colors = [(255 / 255, 160 / 255, 14 / 255), (181 / 255, 87 / 255, 181 / 255)]
    colors_conv = [(210 / 255, 0 / 255, 0 / 255), (27 / 255, 32 / 255, 101 / 255)]

    idx = 10
    idx_filter = 1

    filter = filters[:, 0, idx_filter]

    plt.figure(1)
    plt.plot(filter + 0.5, color='gray', label='filter')
    for c in classes:
        c_x_train = x_train[np.where(y_train == c)]
        convolved_filter_1 = new_feed_forward([c_x_train])[0]

        idx_c = int(c) - 1

        plt.plot(c_x_train[idx], color=colors[idx_c], label='class' + str(idx_c) + '-raw')
        plt.plot(convolved_filter_1[idx, :, idx_filter], color=colors_conv[idx_c], label='class' + str(idx_c) + '-conv')
        plt.legend()

    plt.savefig(root_dir + 'convolution-' + dataset_name + '.pdf')

    return 1


def viz_perf_themes(root_dir, df):
    df_themes = df.copy()
    themes_index = []
    # add the themes
    for dataset_name in df.index:
        themes_index.append(utils.constants.dataset_types[dataset_name])

    themes_index = np.array(themes_index)
    themes, themes_counts = np.unique(themes_index, return_counts=True)
    df_themes.index = themes_index
    df_themes = df_themes.rank(axis=1, method='min', ascending=False)
    df_themes = df_themes.where(df_themes.values == 1)
    df_themes = df_themes.groupby(level=0).sum(axis=1)
    df_themes['#'] = themes_counts

    for classifier in CLASSIFIERS:
        df_themes[classifier] = df_themes[classifier] / df_themes['#'] * 100
    df_themes = df_themes.round(decimals=1)
    df_themes.to_csv(root_dir + 'tab-perf-theme.csv')


def viz_perf_train_size(root_dir, df):
    df_size = df.copy()
    train_sizes = []
    datasets_dict_ucr = read_all_datasets(root_dir, archive_name='UCR_TS_Archive_2015')
    datasets_dict_mts = read_all_datasets(root_dir, archive_name='mts_archive')
    datasets_dict = dict(datasets_dict_ucr, **datasets_dict_mts)

    for dataset_name in df.index:
        train_size = len(datasets_dict[dataset_name][0])
        train_sizes.append(train_size)

    train_sizes = np.array(train_sizes)
    bins = np.array([0, 100, 400, 800, 99999])
    train_size_index = np.digitize(train_sizes, bins)
    train_size_index = bins[train_size_index]

    df_size.index = train_size_index
    df_size = df_size.rank(axis=1, method='min', ascending=False)
    df_size = df_size.groupby(level=0, axis=0).mean()
    df_size = df_size.round(decimals=2)

    print(df_size.to_string())
    df_size.to_csv(root_dir + 'tab-perf-train-size.csv')


def viz_perf_classes(root_dir, df):
    df_classes = df.copy()
    class_numbers = []
    datasets_dict_ucr = read_all_datasets(root_dir, archive_name='UCR_TS_Archive_2015')
    datasets_dict_mts = read_all_datasets(root_dir, archive_name='mts_archive')
    datasets_dict = dict(datasets_dict_ucr, **datasets_dict_mts)

    for dataset_name in df.index:
        train_size = len(np.unique(datasets_dict[dataset_name][1]))
        class_numbers.append(train_size)

    class_numbers = np.array(class_numbers)
    bins = np.array([0, 3, 4, 6, 8, 13, 9999])
    class_numbers_index = np.digitize(class_numbers, bins)
    class_numbers_index = bins[class_numbers_index]

    df_classes.index = class_numbers_index
    df_classes = df_classes.rank(axis=1, method='min', ascending=False)
    df_classes = df_classes.groupby(level=0, axis=0).mean()
    df_classes = df_classes.round(decimals=2)

    print(df_classes.to_string())
    df_classes.to_csv(root_dir + 'tab-perf-classes.csv')


def viz_perf_length(root_dir, df):
    df_lengths = df.copy()
    lengths = []
    datasets_dict_ucr = read_all_datasets(root_dir, archive_name='UCR_TS_Archive_2015')
    datasets_dict_mts = read_all_datasets(root_dir, archive_name='mts_archive')
    datasets_dict = dict(datasets_dict_ucr, **datasets_dict_mts)

    for dataset_name in df.index:
        length = datasets_dict[dataset_name][0].shape[1]
        lengths.append(length)

    lengths = np.array(lengths)
    bins = np.array([0, 81, 251, 451, 700, 1001, 9999])
    lengths_index = np.digitize(lengths, bins)
    lengths_index = bins[lengths_index]

    df_lengths.index = lengths_index
    df_lengths = df_lengths.rank(axis=1, method='min', ascending=False)
    df_lengths = df_lengths.groupby(level=0, axis=0).mean()
    df_lengths = df_lengths.round(decimals=2)

    print(df_lengths.to_string())
    df_lengths.to_csv(root_dir + 'tab-perf-lengths.csv')


def viz_plot(root_dir, df):
    df_lengths = df.copy()
    lengths = []
    datasets_dict_ucr = read_all_datasets(root_dir, archive_name='UCR_TS_Archive_2015')
    datasets_dict_mts = read_all_datasets(root_dir, archive_name='mts_archive')
    datasets_dict = dict(datasets_dict_ucr, **datasets_dict_mts)

    for dataset_name in df.index:
        length = datasets_dict[dataset_name][0].shape[1]
        lengths.append(length)

    lengths_index = np.array(lengths)

    df_lengths.index = lengths_index

    plt.scatter(x=df_lengths['fcn'], y=df_lengths['resnet'])
    plt.ylim(ymin=0, ymax=1.05)
    plt.xlim(xmin=0, xmax=1.05)
    # df_lengths['fcn']
    plt.savefig(root_dir + 'plot.pdf')


def viz_for_survey_paper(root_dir, filename='results-ucr-mts.csv'):
    df = pd.read_csv(root_dir + filename, index_col=0)
    df = df.T
    df = df.round(decimals=2)

    # get table performance per themes
    # viz_perf_themes(root_dir,df)

    # get table performance per train size
    # viz_perf_train_size(root_dir,df)

    # get table performance per classes
    # viz_perf_classes(root_dir,df)

    # get table performance per length
    # viz_perf_length(root_dir,df)

    # get plot
    viz_plot(root_dir, df)


def viz_cam(root_dir):
    import tensorflow.keras as keras
    import sklearn
    classifier = 'resnet'
    archive_name = 'UCRArchive_2018'
    dataset_name = 'GunPoint'

    if dataset_name == 'Gun_Point':
        save_name = 'GunPoint'
    else:
        save_name = dataset_name
    max_length = 2000
    datasets_dict = read_dataset(root_dir, archive_name, dataset_name)

    x_train = datasets_dict[dataset_name][0]
    y_train = datasets_dict[dataset_name][1]
    y_test = datasets_dict[dataset_name][3]

    # transform to binary labels
    enc = sklearn.preprocessing.OneHotEncoder()
    enc.fit(np.concatenate((y_train, y_test), axis=0).reshape(-1, 1))
    y_train_binary = enc.transform(y_train.reshape(-1, 1)).toarray()

    x_train = x_train.reshape(x_train.shape[0], x_train.shape[1], 1)

    model = keras.models.load_model(
        root_dir + 'results/' + classifier + '/' + archive_name + '/' + dataset_name + '/best_model.hdf5')

    # filters
    w_k_c = model.layers[-1].get_weights()[0]  # weights for each filter k for each class c

    # the same input
    new_input_layer = model.inputs
    # output is both the original as well as the before last layer
    new_output_layer = [model.layers[-3].output, model.layers[-1].output]

    new_feed_forward = keras.backend.function(new_input_layer, new_output_layer)

    classes = np.unique(y_train)

    for c in classes:
        plt.figure()
        count = 0
        c_x_train = x_train[np.where(y_train == c)]
        for ts in c_x_train:
            ts = ts.reshape(1, -1, 1)
            [conv_out, predicted] = new_feed_forward([ts])
            pred_label = np.argmax(predicted)
            orig_label = np.argmax(enc.transform([[c]]))
            if pred_label == orig_label:
                cas = np.zeros(dtype=np.float, shape=(conv_out.shape[1]))
                for k, w in enumerate(w_k_c[:, orig_label]):
                    cas += w * conv_out[0, :, k]

                minimum = np.min(cas)

                cas = cas - minimum

                cas = cas / max(cas)
                cas = cas * 100

                x = np.linspace(0, ts.shape[1] - 1, max_length, endpoint=True)
                # linear interpolation to smooth
                f = interp1d(range(ts.shape[1]), ts[0, :, 0])
                y = f(x)
                # if (y < -2.2).any():
                #     continue
                f = interp1d(range(ts.shape[1]), cas)
                cas = f(x).astype(int)
                plt.scatter(x=x, y=y, c=cas, cmap='jet', marker='.', s=2, vmin=0, vmax=100, linewidths=0.0)
                if dataset_name == 'Gun_Point':
                    if c == 1:
                        plt.yticks([-1.0, 0.0, 1.0, 2.0])
                    else:
                        plt.yticks([-2, -1.0, 0.0, 1.0, 2.0])
                count += 1

        cbar = plt.colorbar()
        # cbar.ax.set_yticklabels([100,75,50,25,0])
        plt.savefig(root_dir + '/temp/' + classifier + '-cam-' + save_name + '-class-' + str(int(c)) + '.png',
                    bbox_inches='tight', dpi=1080)
